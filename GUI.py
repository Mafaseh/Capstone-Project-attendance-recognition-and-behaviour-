
import datetime
from datetime import datetime, time
import cv2
import numpy as np
import face_recognition
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from fer import FER
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os
from PIL import ImageTk, Image, ImageFilter, ImageEnhance
from datetime import datetime
import math


# Create the GUI window
window = tk.Tk()
window.geometry("850x650")

# Set window title
window.title("SeeYou")

# Set the window icon
window.iconbitmap("Icon.ico")

# Function to start the attendance system
def start_attendance_system():
    # Define the list of emotions to detect
    emotion_labels = ['happy', 'sad', 'neutral', 'angry']
    path = "C:\\Users\\nasse\\Desktop\\Face-Recognition-master\\Images"
    images = []
    classNames = []
    myList = os.listdir(path)
    print(myList)
    for cl in myList:
        curImg = cv2.imread(f'{path}/{cl}')
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])
    print(classNames)

    def findEncodings(images):
        encodeList = []
        for img in images:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encode = face_recognition.face_encodings(img)[0]
            encodeList.append(encode)
        return encodeList

    def markAttendance(name, action, timestamp=None):
        now = datetime.now()
        dt_string = now.strftime('%Y-%m-%d %H:%M:%S')
        if timestamp:
            dt_string = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        attendance_data = [name, dt_string, action]
        filename = 'Attendance.xlsx'

        # Check if the attendance file already exists
        if os.path.exists(filename):
            wb = load_workbook(filename)
            if wb.sheetnames:
                sheet = wb[wb.sheetnames[0]]
            else:
                sheet = wb.create_sheet(title='Attendance')
                sheet.column_dimensions[get_column_letter(2)].width = 17  # Set the width of column B to 17
                sheet.append(['Name', 'Timestamp', 'Action'])
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'Attendance'
            sheet.column_dimensions[get_column_letter(2)].width = 17  # Set the width of column B to 17
            sheet.append(['Name', 'Timestamp', 'Action'])

        sheet.append(attendance_data)
        wb.save(filename)
        print(f'{name} {action}ed the class.')

    def initializeAttendance(classNames):
        filename = 'Attendance.xlsx'

        if os.path.exists(filename):
            return  # Attendance file already exists

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Attendance'
        sheet.column_dimensions[get_column_letter(2)].width = 17  # Set the width of column B to 17
        sheet.append(['Name', 'Timestamp', 'Action'])

        for name in classNames:
            attendance_data = [name, '', 'Absent']
            sheet.append(attendance_data)

        wb.save(filename)
        print('Attendance initialized.')

    calculate_texture_variance = lambda image: np.var(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)) if image.size else 0

    encodeListKnown = findEncodings(images)
    print('Encoding Complete')

    emotion_detector = FER()

    detectedNames = {}  # Dictionary to store the detected face names and their attendance status

    # Initialize the two camera instances
    emotion_camera = cv2.VideoCapture(0)  # Camera for emotion detection
    attendance_camera = cv2.VideoCapture(1)  # Camera for face recognition and attendance

    # Initialize the figure and axes for the live pie chart
    fig = plt.figure(figsize=(5, 5))
    ax = fig.add_subplot(111)
    plt.axis('off')

    # Variables to store emotion data
    emotion_labels = []
    emotion_values = []

    # Create initial empty pie chart
    patches, texts, autotexts = ax.pie(emotion_values, labels=emotion_labels, autopct='%1.1f%%')
    ax.axis('equal')

    # Update function for the live pie chart
    def update_pie_chart(frame):
        ax.clear()
        patches, texts, autotexts = ax.pie(emotion_values, labels=emotion_labels, autopct='%1.1f%%')
        ax.axis('equal')
        for text in texts:
            text.set_fontsize(12)
        for autotext in autotexts:
            autotext.set_fontsize(12)
            autotext.set_color('white')

    # Animation function for updating the live pie chart
    ani = animation.FuncAnimation(fig, update_pie_chart, frames=1, interval=200, repeat=True)

    # Create a new window for the live pie chart
    plt.show(block=False)

    # Initialize the attendance file
    initializeAttendance(classNames)

    while True:
        # Emotion detection
        success, img_emotion = emotion_camera.read()
        img_emotion = cv2.flip(img_emotion, 1)

        # Face recognition and attendance
        success, img_attendance = attendance_camera.read()
        img_attendance = cv2.flip(img_attendance, 1)
        imgS = cv2.resize(img_attendance, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        newDetectedNames = []  # Temporary list to store newly detected names

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                # Calculate the variance of the face texture
                face_image = img_attendance[y1:y2, x1:x2]

                face_variance = calculate_texture_variance(face_image)
                variance_threshold = 2150  # Adjust this threshold based on your environment

                newDetectedNames.append(name)
                if face_variance > variance_threshold:
                    cv2.rectangle(img_attendance, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    cv2.rectangle(img_attendance, (x1, y2 - 35), (x2, y2), (0, 0, 255), cv2.FILLED)
                    cv2.putText(img_attendance, "Spoof Face", (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1,
                                (255, 255, 255), 2)
                else:
                    cv2.rectangle(img_attendance, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img_attendance, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img_attendance, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    if name not in detectedNames or (name in detectedNames and not detectedNames[name]):
                        current_time = datetime.now().time()
                        if time(8, 30) <= current_time <= time(8, 45) or time(11, 0) <= current_time <= time(11,
                                                                                                             15) or time(
                                13, 30) <= current_time <= time(13, 45):
                            markAttendance(name, 'Present', current_time)
                        else:
                            markAttendance(name, 'Absent')
                        detectedNames[name] = True

            cv2.imshow('Attendance', img_attendance)
            cv2.waitKey(1)

        # Perform emotion detection on the current frame
        results = emotion_detector.detect_emotions(img_emotion)

        # Clear the emotion data for the current frame
        emotion_labels.clear()
        emotion_values.clear()

        for face in results:
            x, y, w, h = face['box']
            emotions = face['emotions']
            emotion_label = max(emotions, key=emotions.get)
            emotion_value = emotions[emotion_label]
            emotion_labels.append(emotion_label)
            emotion_values.append(emotion_value)

            # Draw the rectangle around the face and display the emotion label
            cv2.rectangle(img_emotion, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.putText(img_emotion, emotion_label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        # Update the live pie chart
        update_pie_chart(0)
        plt.pause(0.01)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the camera instances
    emotion_camera.release()
    attendance_camera.release()

    # Destroy all OpenCV windows
    cv2.destroyAllWindows()

    # Show the final pie chart
    plt.show()
    print("Attendance system started.")

# Initialize the attendance data list
attendance_data = []


# Function to view attendance records and extract Excel file
def view_attendance_records():
        try:
            os.startfile("Attendance.xlsx")
        except FileNotFoundError:
            messagebox.showerror("Error", "Attendance records file not found.")


# Load the image
image_path = "SeeYou.png"  # Replace "path/to/image.png" with the actual path to your image file
original_image = Image.open(image_path)

# Resize the image
resized_image = original_image.resize((200, 150))  # Specify the desired width and height

# Convert the resized image to Tkinter PhotoImage
image = ImageTk.PhotoImage(resized_image)

# Create a label for the image
image_label = tk.Label(window, image=image)
# Set the position of the image using the place() method
image_label.place(x=325, y=-10)  # Adjust the x and y coordinates as needed

# Language translations
language_translations = {
    "English": {
        "language_message": "Language changed to English.",
        "title": "SeeYou",
        "subtitle": "Student Attendance & Behaviour System",
        "start_button": "Start",
        "settings_button": "Settings",
        "about_button": "About",
        "change_theme_button": "Change Theme",
        "view_attendance_records_button": "View Attendance Records",
        "change_language_button": "Change Language"
    },
    "French": {
        "language_message": "Langue changée en français.",
        "title": "SeeYou",
        "subtitle": "Système de présence et de comportement des étudiants",
        "start_button": "Démarrer",
        "settings_button": "Paramètres",
        "about_button": "À propos",
        "change_theme_button": "Changer le thème",
        "view_attendance_records_button": "Voir les enregistrements de présence",
        "change_language_button": "Changer la langue"
    },
    "Arabic": {
        "language_message": "تم تغيير اللغة إلى العربية.",
        "title": "SeeYou",
        "subtitle": "نظام حضور وسلوك الطلاب",
        "start_button": "بدء",
        "settings_button": "الإعدادات",
        "about_button": "حول",
        "change_theme_button": "تغيير السمة",
        "view_attendance_records_button": "عرض سجلات الحضور",
        "change_language_button": "تغيير اللغة"
    }
}

# Custom title design
title_label = tk.Label(window, font=("Times New Roman", 24, "bold italic"), fg="dark blue")
title_label.pack(pady=70)
title_label.place(x=370, y=130)




# Create a global variable for the current theme
current_theme = "default"

# Load the original background image
original_background_image = Image.open("Background.png")

# Function to change the theme
def change_theme():
    messagebox.showinfo("Change Theme", "Theme changed.")
    global current_theme, original_background_image

    # Check the current theme and toggle to the opposite theme
    if current_theme == "default":
        # Darken the background image
        background_image = original_background_image.copy()
        background_image = darken_image(background_image)
        current_theme = "dark"
    else:
        background_image = original_background_image.copy()
        current_theme = "default"

    # Convert the background image to Tkinter PhotoImage
    background_photo = ImageTk.PhotoImage(background_image)

    # Update the background image on the main window
    background_label.configure(image=background_photo)
    background_label.image = background_photo  # Store a reference to the image

# Function to darken an image
def darken_image(image):
    # Create an enhancer object for the image
    enhancer = ImageEnhance.Brightness(image)
    # Decrease the brightness by a factor of 0.8 to darken the image
    darkened_image = enhancer.enhance(0.2)
    return darkened_image

# Create a label for the background image
background_label = tk.Label(window)
background_label.place(x=0, y=0)

# Load the original background image
original_background_image = Image.open("Background.png")

# Convert the background image to Tkinter PhotoImage
background_photo = ImageTk.PhotoImage(original_background_image)

# Set the initial background image on the label
background_label.configure(image=background_photo)
background_label.image = background_photo  # Store a reference to the image


# Function to change the language
def change_language():
    global current_language

    # Create a language selection dialog
    language_dialog = tk.Toplevel(window)
    language_dialog.title("Select Language")
    language_dialog.geometry("300x200")
    language_dialog.iconbitmap("Settings.ico")

    # Language selection label
    language_label = tk.Label(language_dialog, text="Select a language:", font=("Times New Roman", 12, "bold"))
    language_label.pack(pady=10)

    # Language selection buttons
    def set_language(language):
        global current_language
        current_language = language
        messagebox.showinfo("Language Change", language_translations[current_language]["language_message"])
        update_language()  # Update the language translations
        language_dialog.destroy()  # Close the language selection dialog

    english_button = tk.Button(language_dialog, text="English", command=lambda: set_language("English"), bg="dark blue", fg="#FF1493",
                         font=("Times New Roman", 12), width=5, height=1, activebackground="light blue", activeforeground="white")
    english_button.pack(pady=5)

    french_button = tk.Button(language_dialog, text="French", command=lambda: set_language("French"), bg="dark blue", fg="#FF1493",
                         font=("Times New Roman", 12), width=5, height=1, activebackground="light blue", activeforeground="white")
    french_button.pack(pady=5)

    arabic_button = tk.Button(language_dialog, text="Arabic", command=lambda: set_language("Arabic"), bg="dark blue", fg="#FF1493",
                         font=("Times New Roman", 12), width=5, height=1, activebackground="light blue", activeforeground="white")
    arabic_button.pack(pady=5)

    # Run the language selection dialog
    language_dialog.mainloop()

# Create a settings window
settings_window = tk.Toplevel(window)
settings_window.title("Settings")
settings_window.geometry("300x200")
settings_window.iconbitmap("Icon.ico")
settings_window.withdraw()  # Hide the settings window initially


# Function to open the settings window
def open_settings_window():
    settings_window.deiconify()  # Show the settings window


# Function to close the settings window
def close_settings_window():
    settings_window.withdraw()  # Hide the settings window

# Variable to keep track of the current language
current_language = "English"


# Subtitle label
subtitle_label = tk.Label(window, font=("Times New Roman", 12), fg="#FF1493")
subtitle_label.place(x=680, y=285)

# Start button
start_button = tk.Button(window, command=start_attendance_system, bg="dark blue", fg="#FF1493",
                         font=("Times New Roman", 15, "bold"), width=10, height=2, activebackground="light blue", activeforeground="white")
start_button.place(x=750, y=340)

# Settings button
settings_button = tk.Button(window,text=language_translations[current_language]["settings_button"],
                            command=open_settings_window, bg="dark blue", fg="#FF1493", font=("Times New Roman", 15, "bold"), width=10,
                                height=2, activebackground="light blue", activeforeground="white")
settings_button.place(x=750, y=420)

# Function to display the system description
def display_about():
    # Create a new window or frame for the about information
    about_window = tk.Toplevel(window)
    about_window.title("About")
    # Set the about window icon
    about_window.iconbitmap("About.ico")

    about_text = "This is a student attendance system and behavior tracking system using facial recognition technology. " \
                 "The system requires students to scan their faces through a camera, which is then matched with a database of all registered students. " \
                 "If a match is found, the system marks the attendance and saves it in an XLSX file. " \
                 "Additionally, this system has the ability to detect students' emotions and identify fake attendance attempts."

    about_text_widget = tk.Text(about_window, wrap="word", font=("Times New Roman", 12), height=4, spacing1= 3)
    about_text_widget.insert("1.0", about_text)
    about_text_widget.configure(state="disabled")
    about_text_widget.pack(padx=20, pady=20)


# About button
about_button = tk.Button(window, command=display_about, bg="dark blue", fg="#FF1493",
                         font=("Times New Roman", 15, "bold"), width=10, height=2, activebackground="light blue", activeforeground="white")
about_button.place(x=750, y=500)


change_theme_button = tk.Button(settings_window, text=language_translations[current_language]["change_theme_button"],
                                command=change_theme, bg="dark blue", fg="#FF1493", font=("Times New Roman", 12, "bold"), width=11, height=1, activebackground="light blue", activeforeground="white")
change_theme_button.pack(pady=5)


# View attendance records button
view_attendance_records_button = tk.Button(settings_window,
                                           text=language_translations[current_language]["view_attendance_records_button"],
                                           command=view_attendance_records,  bg="dark blue", fg="#FF1493", font=("Times New Roman", 12, "bold"), width=19, height=1, activebackground="light blue", activeforeground="white")
view_attendance_records_button.pack(pady=5)

# Change language button
change_language_button = tk.Button(settings_window,
                                   text=language_translations[current_language]["change_language_button"],
                                   command=change_language,  bg="dark blue", fg="#FF1493", font=("Times New Roman", 12, "bold"), width=13, height=1, activebackground="light blue", activeforeground="white")
change_language_button.pack(pady=5)


# Current date and time label
current_datetime_label = tk.Label(window, fg="#FF1493", font=("Times New Roman", 12, "bold"))
current_datetime_label.place(x=70, y=15)


# Load the clock image
clock_image = Image.open("Analog.png")  # Replace "clock_image.png" with your own clock image file
clock_image = clock_image.resize((200, 200))  # Resize the image to fit the canvas

# Convert the clock image to Tkinter PhotoImage
clock_photo = ImageTk.PhotoImage(clock_image)

# Create a canvas and set the clock image as its background
canvas = tk.Canvas(window, width=200, height=200)
canvas.place(x=10, y=70)
canvas.create_image(100, 100, image=clock_photo)

# Function to update the analog clock
def update_clock():
    now = datetime.now()
    hour = now.hour
    minute = now.minute
    second = now.second

    # Convert the time to degrees
    hour_angle = (hour % 12) * 30 + minute / 2
    minute_angle = minute * 6
    second_angle = second * 6

    # Draw the clock hands
    canvas.delete("clock_hands")
    canvas.create_line(100, 100, 100 + 60 * math.cos(math.radians(hour_angle - 90)),
                       100 + 60 * math.sin(math.radians(hour_angle - 90)), width=3, fill="black", tags="clock_hands")
    canvas.create_line(100, 100, 100 + 80 * math.cos(math.radians(minute_angle - 90)),
                       100 + 80 * math.sin(math.radians(minute_angle - 90)), width=2, fill="black", tags="clock_hands")
    canvas.create_line(100, 100, 100 + 90 * math.cos(math.radians(second_angle - 90)),
                       100 + 90 * math.sin(math.radians(second_angle - 90)), width=1, fill="red", tags="clock_hands")

    window.after(1000, update_clock)  # Update every 1 second


# Run the update_clock function to start updating the analog clock
update_clock()

# Function to update the current date and time label
def update_datetime():
    now = datetime.now()
    current_datetime = now.strftime("%d-%m-%Y \n %H:%M:%S")
    current_datetime_label.config(text=current_datetime)
    window.after(1000, update_datetime)  # Update every 1 second


# Run the update_datetime function to start updating the current date and time label
update_datetime()

# Function to update the language translations
def update_language():
    global current_language
    title_label.config(text=language_translations[current_language]["title"])
    subtitle_label.config(text=language_translations[current_language]["subtitle"])
    start_button.config(text=language_translations[current_language]["start_button"])
    settings_button.config(text=language_translations[current_language]["settings_button"])
    about_button.config(text=language_translations[current_language]["about_button"])


# Run the update_language function to set the initial language translations
update_language()

# Run the GUI event loop
window.mainloop()